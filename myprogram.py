import sys
from PyQt5 import QtGui
from PyQt5 import QtWidgets,QtCore
import idna
import numpy as np
import openpyxl
import pandas as pd
import datetime
my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
my_blue = openpyxl.styles.colors.Color(rgb='000000FF')
my_green = openpyxl.styles.colors.Color(rgb='0000FF00')
my_yellow = openpyxl.styles.colors.Color(rgb='00FFFF00')

my_fill_red = openpyxl.styles.fills.PatternFill(patternType='mediumGray', fgColor=my_red)
my_fill_blue = openpyxl.styles.fills.PatternFill(patternType='mediumGray', fgColor=my_blue)
my_fill_green = openpyxl.styles.fills.PatternFill(patternType='mediumGray', fgColor=my_green)
my_fill_yellow = openpyxl.styles.fills.PatternFill(patternType='mediumGray', fgColor=my_yellow)

wb = openpyxl.load_workbook('Book2.xlsx')
sheet= wb.get_active_sheet()
 



class Window(QtWidgets.QMainWindow):
    __month_ind='Z'
    __resident_cash=str('none')
    __resident_imps=str('none')
    __amount_cash=0.0
    __reference_number=" "
    __mode=[]
    __amt=[]
    __amt_imps=0.0
    __name='-Select-'             #pdf file name
    def __init__(self):
        super(Window,self).__init__()
        self.setGeometry(50,50,650,850)
        self.setWindowTitle("ANANYA APPARTMENTS")
        self.setWindowIcon(QtGui.QIcon("logoV.jpg"))
        self.home()
    def home(self):

        self.title_label = QtWidgets.QLabel(self)
        self.title_label.move(100,30)
        #label = QLabel(self)
        self.pixmap = QtGui.QPixmap('logo.jpeg')
        self.title_label.setPixmap(self.pixmap)

        # Optional, resize window to image size
        self.title_label.resize(self.pixmap.width(),self.pixmap.height())

        self.cash_label = QtWidgets.QLabel( "<P><b><i><FONT COLOR='#0000ff' FONT SIZE = 4>For Cash Payments:</i></b></P></br>", self)
        self.cash_label.resize(200,30)
        self.cash_label.move(50,500)

        self.cash_label = QtWidgets.QLabel( "<P><b><i><FONT COLOR='#ff0000' FONT SIZE = 4>Update using Reference Number:</i></b></P></br>", self)
        self.cash_label.resize(300,30)
        self.cash_label.move(50,650)

        
        self.fileChoice = QtWidgets.QLabel("Select File", self)
        self.filename = QtWidgets.QLabel(self.__name, self)
        btn_file=QtWidgets.QPushButton("Select File",self)
        btn_file.clicked.connect(self.file_code)
        btn_file.move(150,450)
        btn_file.resize(100,32)
        self.fileChoice.move(50,450)
        self.filename.move(275,450)

        btn_upload=QtWidgets.QPushButton("Upload File",self)
        btn_upload.clicked.connect(self.upload_code)
        btn_upload.move(430,450)
        btn_upload.resize(100,32)
        self.upload_label = QtWidgets.QLabel( " ", self)
        self.upload_label.move(540,450)

        

        btn_amt=QtWidgets.QPushButton("Amount",self)
        btn_amt.clicked.connect(self.amt_code)
        btn_amt.move(150,600)
        btn_amt.resize(100,32)
        self.amount_label = QtWidgets.QLabel( 'Rs.' + str(self.__amount_cash), self)
        self.amount_label.move(260,600)
        

        btn_cash=QtWidgets.QPushButton("Update Cash",self)
        btn_cash.clicked.connect(self.cash_code)
        btn_cash.move(430,550)
        btn_cash.resize(100,32)
        self.update_label = QtWidgets.QLabel( " ", self)
        self.update_label.move(540,550)

        self.monthChoice = QtWidgets.QLabel("Select Month", self)

        comboBox = QtWidgets.QComboBox(self)
        comboBox.addItem("-Select-")
        comboBox.addItem("Jan")
        comboBox.addItem("Feb")
        comboBox.addItem("Mar")
        comboBox.addItem("Apr")
        comboBox.addItem("May")
        comboBox.addItem("Jun")
        comboBox.addItem("Jul")
        comboBox.addItem("Aug")
        comboBox.addItem("Sep")
        comboBox.addItem("Oct")
        comboBox.addItem("Nov")
        comboBox.addItem("Dec")
        comboBox.move(150, 400)
        
        self.monthChoice.move(50,400)
        comboBox.activated[str].connect(self.month_choice)


        self.residentChoice = QtWidgets.QLabel("Select Resident", self)

        comboBox1 = QtWidgets.QComboBox(self)
        comboBox1.addItem("-Select-")
        for i in range(2,sheet.max_row+1):
            
            data=sheet['B'+str(i)].value
            if(data!=None):
                comboBox1.addItem(sheet['A'+str(i)].value + "-"+data)
                
        
        comboBox1.move(150, 550)
        comboBox1.resize(200, 34)
        
        self.residentChoice.move(50,550)
        comboBox1.activated[str].connect(self.resident_choice)

        btn_ref=QtWidgets.QPushButton("Reference No.",self)
        btn_ref.clicked.connect(self.ref_code)
        btn_ref.move(150,700)
        btn_ref.resize(100,32)
        self.ref_label = QtWidgets.QLabel( str(self.__reference_number), self)
        self.ref_label.move(260,700)


        
        self.valid_label = QtWidgets.QLabel( "No", self)
        self.valid_label.move(150,750)

        self.residentChoice2 = QtWidgets.QLabel("Select Resident", self)

        comboBox2 = QtWidgets.QComboBox(self)
        comboBox2.addItem("-Select-")
        for i in range(2,sheet.max_row+1):
            
            data=sheet['B'+str(i)].value
            if(data!=None):
                comboBox2.addItem(sheet['A'+str(i)].value + "-"+data)
                
        
        comboBox2.move(150, 800)
        comboBox2.resize(200, 34)
        
        self.residentChoice2.move(50,800)
        comboBox2.activated[str].connect(self.resident_choice_imps)

        btn_update_imps=QtWidgets.QPushButton("Update",self)
        btn_update_imps.clicked.connect(self.upload_imps__code)
        btn_update_imps.move(430,800)
        btn_update_imps.resize(100,32)
        self.update_imps_label = QtWidgets.QLabel( " ", self)
        self.update_imps_label.move(540,800)
        
        self.show()
    def ref_code(self):
        self.__reference_number,ret = QtWidgets.QInputDialog.getText(self,"Reference Number","Enter the Reference Number")
        self.ref_label.setText(str(self.__reference_number))
        k=0
        for j in self.__mode:
            #print(j)
            if(str(j[0])!='nan'):
                
                if(int(j[0])==int(self.__reference_number)):
                    self.__amt_imps=self.__amt[k][0]
                    self.valid_label.setText("Yes  Amt:" + str(self.__amt[k][0])     )           
            k=k+1
        
    def upload_imps__code(self):
        if(self.__name=='-Select-'):
            self.update_imps_label.setText("Upload File")
            return
        if(self.__resident_imps=='none'):
            self.update_imps_label.setText('Select Resident')
            return
        if(self.__amt_imps==0.0):
            self.update_imps_label.setText('Enter Ref No.')
            return
        if(self.__month_ind=='Z'):
            self.update_imps_label.setText('Select Month')
            return
        
        for i in range(2,(sheet.max_row)+1):
            #data=sheet.cell_value(i,1)
            data=sheet['B'+str(i)].value
            #print(data)
            if(data!=None):
                if((data.lower())==self.__resident_imps.lower()):
                    sheet[self.__month_ind+str(i)]=float(self.__amt_imps)
                    sheet[self.__month_ind+str(i)].fill=my_fill_yellow
                    #sheet['P' + str(i)]=(float(sheet['P' + str(i)].value) + float(amount))

        for i in range(3,(sheet.max_row)+1):
             total=0
             for j in range(4,16):
                 if(sheet.cell(row=i,column=j).value is not None):
                     total=float(sheet.cell(row=i,column=j).value)+total
             sheet['P' + str(i)]=total 

        wb.save('Book2.xlsx')
        self.update_imps_label.setText("Updated")

        
        
    def resident_choice_imps(self,resident):
        self.__resident_imps=resident.split('-')[1]
        print(self.__resident_cash)
    def resident_choice(self,resident):
        self.__resident_cash=resident.split('-')[1]
        print(self.__resident_cash)
    def amt_code(self):
        self.__amount_cash,ret = QtWidgets.QInputDialog.getDouble(self,"Cash Amount","Enter the Cash Amount")
        self.amount_label.setText(str(self.__amount_cash))
        
    def cash_code(self):
        if(self.__resident_cash=='none'):
            self.update_label.setText('Select Resident')
            return
        if(self.__amount_cash==0.0):
            self.update_label.setText('Enter Amount')
            return
        if(self.__month_ind=='Z'):
            self.update_label.setText('Select Month')
            return
        
        for i in range(2,(sheet.max_row)+1):
            #data=sheet.cell_value(i,1)
            data=sheet['B'+str(i)].value
            #print(data)
            if(data!=None):
                if((data.lower())==self.__resident_cash.lower()):
                    sheet[self.__month_ind+str(i)]=float(self.__amount_cash)
                    sheet[self.__month_ind+str(i)].fill=my_fill_green
                    #sheet['P' + str(i)]=(float(sheet['P' + str(i)].value) + float(amount))

        for i in range(3,(sheet.max_row)+1):
             total=0
             for j in range(4,16):
                 if(sheet.cell(row=i,column=j).value is not None):
                     total=float(sheet.cell(row=i,column=j).value)+total
             sheet['P' + str(i)]=total 

        wb.save('C:\\Users\\V Vignesh\\Desktop\\project\\python-pdftables-api\\Book2.xlsx')
        self.update_label.setText("Updated")

        
        
        
    def month_choice(self,month):
        
        self.__month_ind='D'
        for i in range(4,15):
            if(  (sheet[(self.__month_ind)+str(1)].value.strftime("%b") ).lower()== month.lower()):
                break;
            else:
                self.__month_ind=ord(self.__month_ind[0])
                self.__month_ind=self.__month_ind+1
                self.__month_ind=chr(self.__month_ind)
                
            
    def file_code(self):
        print("hello")
        #name = QtGui.QFileDialog.getOpenFileName(self, 'Open File')
        self.__name = QtWidgets.QFileDialog.getOpenFileName(self, 'Open File')
        #print(name[0])
        #self.online(name[0])
        self.filename.setText(self.__name[0].split('/')[-1])
    
    
        
    def upload_code(self):
        
        if(self.__name=='-Select-'):
            self.upload_label.setText("Select File")
            return
        if(self.__month_ind=='Z'):
            self.upload_label.setText("Select Month")
            return
            
        #df = tabula.read_pdf(self.__name[0], pages='all')
        df=pd.read_excel(self.__name[0])
        #df.iloc[6,1]='CHEQUE'
        #for i in range(1,df.shape[0]-1):
        #    if(  pd.notna(df.iloc[i,3]) and pd.isna(df.iloc[i-1,0]) and pd.isna(df.iloc[i+1,0]) ):
        #        df.iat[i,2]=df.iat[i-1,2] + df.iat[i+1,2]

        #df=df.dropna(thresh=2)
        df1=(df.loc[df.iloc[:,5].notna()  ] )
        trans=(df1.iloc[:,[1]].values)
        self.__mode=(df1.iloc[:,[2]].values)
        self.__amt=(df1.iloc[:,[5]].values)
        l=len(trans)
        trans1=[]
        for i in range(1,l-1):
            trans1.append(str(trans[i,0]))
        #print(self.__mode)
        for i in range(2,sheet.max_row+1):
            
            data=sheet['B'+str(i)].value
            data1=sheet['R'+str(i)].value
        
            if((data)!=None):
                for j in trans1:
                    k=str(j.lower())
                    if(k.find(data.lower())!=-1):
                        #print(mode[trans1.index(j)])
                        if(j[0:6] == 'By Clg'):
                            color=my_fill_red
                        else:
                            color=my_fill_blue
                        print("found",data)
                        amount=(self.__amt[trans1.index(j)][0])
                        print(amount)
                        sheet[self.__month_ind+str(i)]=float(amount)
                        sheet[self.__month_ind+str(i)].fill=color

                    elif((data1)!=None):
                        if(k.find(data1.lower())!=-1):
                            #print(mode[trans1.index(j)])
                            if(j[0:6] == 'By Clg'):
                                color=my_fill_red
                            else:
                                color=my_fill_blue
                            print("found",data1)
                            amount=(self.__amt[trans1.index(j)][0])
                            print(amount)
                            sheet[self.__month_ind+str(i)]=float(amount)
                            sheet[self.__month_ind+str(i)].fill=color

                        

        for i in range(3,(sheet.max_row)+1):
            total=0
            for j in range(4,16):
                if(sheet.cell(row=i,column=j).value is not None):
                    total=float(sheet.cell(row=i,column=j).value)+total
            sheet['P' + str(i)]=total 

        wb.save('Book2.xlsx')
        self.upload_label.setText("Updated")

        

        
   


def run():
    app=QtWidgets.QApplication([])
    GUI=Window()
    sys.exit(app.exec_())



run()    
